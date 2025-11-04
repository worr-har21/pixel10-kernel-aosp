# SPDX-License-Identifier: GPL-2.0-only

"""Script to parse the IO sheet and generate the linux_drivers/pinctrl/google/pinctrl-google-lga.c file.

The sheet needs to be downloaded from go/<target>-io and should be saved in the same
directory as that of the parser.

Current sheet date: Sep 20, 2023.

Pre-requisite:
Excel parser module. 'pip install openpyxl' if not exists.

Arguments: [OPTIONAL] --out_file: Writes the output to the filename provided
"""

import argparse
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class Pin:
  """Represents a physical pin/pad."""

  def __init__(self, pad_nr, pad_name, pin_functions):
    self.pad_nr = int(pad_nr.split("_")[1])
    self.pad_name = pad_name
    self.pin_functions = pin_functions


class Sswrp:
  """Represents a subsystem wrapper(sswrp)."""

  def __init__(self, name):
    self.name = "_".join(name.split("_")[:-1]).lower()
    self.pins = []
    self.functions = {}
    self.sswrp_data_name = "google_lga_pinctrl_" + self.name
    self.pins_array_name = "google_lga_" + self.name
    self.pin_groups_name = "google_lga_" + self.name + "_groups"

    self.pin_functions_name = "google_lga_" + self.name + "_functions"
    self.functions_enum_name = "google_lga_pinmux_" + self.name + "_functions"

  def print_pinctrl_pins(self):
    print("static const struct pinctrl_pin_desc {}[] = {{".format(
        self.pins_array_name))

    for i in range(0, len(self.pins)):
      print("\tPINCTRL_PIN({}, \"{}\"),".format(i,self.pins[i].pad_name,))
    print("};\n")

  def print_pin_groups(self):
    print("static const struct google_pingroup {}[] = {{".format(
        self.pin_groups_name))
    for i in range(0, len(self.pins)):
      print("\tPIN_GROUP({}, \"{}\"".format(i, self.pins[i].pad_name), end="")
      for j in range(0, len(self.pins[i].pin_functions)):
        print(", {}".format(self.pins[i].pin_functions[j]), end="")
      print("),")
    print("};\n")

  def print_functions_enum(self):
    print("enum {}{{".format(self.functions_enum_name))
    for func in self.functions:
      print("\tgoogle_pinmux_{},".format(func))
    print("};\n")

  def print_function_groups(self):
    for func_name in self.functions:
      print("FUNCTION_GROUPS({}".format(func_name), end="")
      for pin in self.functions[func_name]:
        print(", \"{}\"".format(pin), end="")
      print(");")
    print()

  def print_functions(self):
    print("static const struct google_pin_function {}[] = {{".format(
        self.pin_functions_name))
    for func in self.functions:
      print("\tFUNCTION({}),".format(func))
    print("};")

  def print_max_gpios(self):
    print("\n#define MAX_NR_GPIO_{} {}\n".format(self.name.upper(),
                                               len(self.pins)))

  # pylint: disable=missing-function-docstring
  def print_sswrp_info(self):

    print("static const struct google_pinctrl_soc_sswrp_info {} = {{".format(
        self.sswrp_data_name))

    print("\t.pins = {},".format(self.pins_array_name))
    print("\t.num_pins = ARRAY_SIZE({}),".format(self.pins_array_name))

    print("\t.groups = {},".format(self.pin_groups_name))
    print("\t.num_groups = ARRAY_SIZE({}),".format(self.pin_groups_name))

    print("\t.funcs = {},".format(self.pin_functions_name))
    print("\t.num_funcs = ARRAY_SIZE({}),".format(self.pin_functions_name))

    print("\t.gpio_func = GPIO_FUNC_BIT_POS,")
    print("\t.num_gpios = MAX_NR_GPIO_{}".format(self.name.upper()))
    print("};\n")


class Common:
  """Contains utility functions that are not specific to a particulat SSWRP."""

  def __init__(self, sswrp_list):
    self.sswrp_list = sswrp_list

  # pylint: disable=missing-function-docstring
  def print_compatible(self):

    def print_compatible_single(sswrp):
      compatible_string = "google," + "lga-" + sswrp.name + "-pinctrl"
      compatible_name = compatible_string.replace("_", "-")

      print("\t{{ .compatible = \"{}\",".format(compatible_name))
      print("\t  .data = &{} }},".format(sswrp.sswrp_data_name))

    match_table_name = "google_lga_of_match"
    print("static const struct of_device_id {}[] = {{".format(match_table_name))
    for sswrp in self.sswrp_list:
      print_compatible_single(sswrp)
    print("};")

  def print_google_pins(self):
    max_pins = 0
    for sswrp in self.sswrp_list:
      max_pins = max(max_pins, len(sswrp.pins))
    for i in range(0, max_pins):
      print("GOOGLE_PINS({});".format(i))
    print()

  def print_header(self):
    print("""// SPDX-License-Identifier: GPL-2.0-only
/*
 * Copyright 2023 Google LLC
 */
#include <linux/module.h>
#include <linux/of.h>
#include <linux/pinctrl/pinctrl.h>
#include <linux/platform_device.h>

#include "common.h"
#define google_pinmux__ -1
""")

  def print_footer(self):
    print("""
static int google_lga_pinctrl_probe(struct platform_device *pdev)
{
	return google_pinctrl_probe(pdev, google_lga_of_match);
}

static struct platform_driver google_lga_pinctrl_driver = {
	.driver = {
                .name = "google_lga_pinctrl",
                .of_match_table = google_lga_of_match,
#ifdef CONFIG_PM
                .pm = &google_pinctrl_pm_ops,
#endif
	},
	.probe = google_lga_pinctrl_probe,
	.remove = google_pinctrl_remove,
};

static int __init google_lga_pinctrl_init(void)
{
	return platform_driver_register(&google_lga_pinctrl_driver);
}
arch_initcall(google_lga_pinctrl_init);

static void __exit google_lga_pinctrl_exit(void)
{
	platform_driver_unregister(&google_lga_pinctrl_driver);
}
module_exit(google_lga_pinctrl_exit);

MODULE_AUTHOR("Google LLC");
MODULE_DESCRIPTION("Google LGA PINCTRL Driver");
MODULE_LICENSE("GPL");\n""")


class Parser:
  """Handles the parsing from IO sheet."""

  def __init__(self):
    self.wb = load_workbook("data_sheet.xlsx")
    self.ws = self.wb["IO pads"]
    self.row_count = self.ws.max_row
    # The SSWRPs for which we want to extract the pin information
    self.valid_sswrps = [
        "HSION_BK", "HSIOS_BK", "HSIOS_STBY_BK", "LSIOS_BK", "LSIOE_BK", "LSION_BK", "CPM_BK",
        "AOC_BK", "GDMC_BK", "DPU_BK"
    ]
    # At least one pin of this SSWRP has been found
    self.accessed_sswrps = []
    self.sswrp_list = []
    # List of valid functions for every SSWRP
    self.valid_functions = {
        "hsion": [
            "gpio", "pcie0", "atb0"
        ],
        "hsios": [
            "gpio", "ufs", "sd_data", "sd_cmd", "sd_fbclk", "sd_clk",
            "atb1", "pcie1", "vsync"
        ],
        "hsios_stby": [
            "clkbuf", "ufs"
        ],
        "lsios": [
            "gpio", "mclk", "debug_mux", "cli0", "qspi0", "cli1", "pwm", "cli2",
            "cli3", "vsync", "cli5", "cli4", "pre_ocp_gpu", "soft_pre_ocp_gpu",
            "cli6", "camera_mute"
        ],
        "lsioe": [
            "gpio", "cli7", "cli8", "cli13", "cli9", "cli10", "cli11", "cli12",
            "cli14", "spmi0"
        ],
        "lsion": [
            "gpio", "cli15", "cli16"
        ],
        "cpm": [
            "int", "clkbuf", "pwrok", "pwr", "reset", "vdroop", "spmi", "uart0",
            "m0_uart", "stby", "pre", "irq", "clkout", "xtal", "boot", "otp_emu",
            "xom", "spi", "i2c", "src"
        ],
        "aoc": [
            "int_gpio", "pdm", "uart0", "uart1", "uart2", "uart3", "uart4",
            "uart5", "i2s", "spi0", "spi1", "spi2", "spi3", "spi4", "spi5",
            "out_mclk", "i3c0", "i3c1", "i3c2", "i3c3", "i3c4", "i2c0", "i2c1",
            "i2c2", "i2c3", "i2c4", "i3c5", "pdm0_flckr", "elaoutput",
            "sdwire0", "sdwire1", "cpm_uart", "tdm0", "tdm1", "fiq", "pdm1",
            "pdm2", "pdm3", "pdm4", "cpm_eint", "cli_int_0", "cli_int_1",
            "cli_int_2", "cli_int_3", "cli_int_4"
        ],
        "gdmc": [
            "int", "uart", "debug", "jtag", "cti", "dbgsel"
        ],
        "dpu": [
            "gpio", "te"
        ]
    }

  def get_short_fname(self, long_fname, sswrp_name):
    """Eliminates the pin information from the function names.

    The function names in the IO sheet contain pin information in them, this
    function removes it.
    eg. Converts lsios_cli3_pin0 to lsios_cli3.

    Args:
        long_fname: The long function name taken from the IO sheet, eg.
          lsios_cli3_pin0.
        sswrp_name: The name of the sswrp to which this pin belongs.

    Returns:
        Short name of the function, eg. lsios_cli3.
    """
    short_fnames = []
    for func in self.valid_functions[sswrp_name]:
      if func in long_fname:
        short_fnames.append(func)
    if not short_fnames:
      sys.exit(
          "No matching short function name present for {}".format(long_fname))
    elif len(short_fnames) > 1:
      # Choosing the longest matching short name
      short_fnames.sort(key=lambda x: (-len(x), x))
    return sswrp_name + "_" + short_fnames[0]

  def get_functions(self, row, pad_name):
    """Parses the functions for the pin on the nth row.

    Parses and stores the supported functions for the pin on the
    row'th row. It also creates an alternate mapping between the
    function name and the pins that can support it.

    Args:
        row: The query row in the IO sheet.
        pad_name: String that contains the name of the pin on the row'th row.

    Returns:
        An array that contains all the functions supported
        by the pin on the given row.
    """
    pin_functions = []
    func0_col = 40
    max_nr_funcs = 9

    for i in range(0, max_nr_funcs):
      function_col = get_column_letter(func0_col + i * 2)
      function_cell = function_col + str(row)

      if self.ws[function_cell].value is not None and str(self.ws[function_cell].value.lower()) != "no" :
        long_fname = self.ws[function_cell].value.lower()
        short_fname = self.get_short_fname(long_fname, self.sswrp_list[-1].name)
        pin_functions.append(short_fname)

        # For every function, populating the pins that support it
        if short_fname not in self.sswrp_list[-1].functions:
          self.sswrp_list[-1].functions[short_fname] = []
        self.sswrp_list[-1].functions[short_fname].append(pad_name)
      else:
        pin_functions.append("_")
    return pin_functions

  # Gets executed when we access a valid SSWRP for the first time
  def check_and_append_sswrp(self, sswrp_name):
    if sswrp_name not in self.accessed_sswrps:
      self.sswrp_list.append(Sswrp(sswrp_name))
      self.accessed_sswrps.append(sswrp_name)

  # A valid pin belongs to a valid SSWRP and has a non empty pad number field
  def is_pin_valid(self, pad_nr_cell, sswrp_name):
    return (sswrp_name in self.valid_sswrps) and (self.ws[pad_nr_cell].value
                                                  is not None)

  def parse_data(self):
    """The main function used to parse data."""
    for row in range(1, self.row_count + 1):
      sswrp_cell = "J" + str(row)
      sswrp_name = self.ws[sswrp_cell].value
      pad_nr_cell = "B" + str(row)
      pad_name_cell = "C" + str(row)
      pad_name = self.ws[pad_name_cell].value

      if self.is_pin_valid(pad_nr_cell, sswrp_name):
        self.check_and_append_sswrp(sswrp_name)
        self.sswrp_list[-1].pins.append(
            Pin(self.ws[pad_nr_cell].value, pad_name,
                self.get_functions(row, pad_name)))

    return self.sswrp_list


def main():

  arg_parser = argparse.ArgumentParser()
  arg_parser.add_argument("-f", "--out_file", help="Writes output to file")
  args = arg_parser.parse_args()
  if args.out_file:
    sys.stdout = open(args.out_file, "w")

  parser = Parser()
  sswrp_list = parser.parse_data()

  common = Common(sswrp_list)
  common.print_header()
  common.print_google_pins()

  for sswrp in sswrp_list:
    sswrp.print_pinctrl_pins()
    sswrp.print_functions_enum()
    sswrp.print_function_groups()
    sswrp.print_pin_groups()

    sswrp.print_functions()
    sswrp.print_max_gpios()
    sswrp.print_sswrp_info()
    print()

  common.print_compatible()
  common.print_footer()


if __name__ == "__main__":
  main()
